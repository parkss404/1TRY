from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter

def write_excel_template(filename,sheetname,listdata):
    excel_file=openpyxl.Workbook()
    excel_sheet=excel_file.active
    excel_sheet.cell(1,1,'가게이름')
    excel_sheet.cell(1,2,'별점')
    excel_sheet.cell(1,3,'리뷰수')
    excel_sheet.cell(1,4,'배달시간')

    if sheetname != '':
        excel_sheet.title=sheetname

    for item in listdata:
        excel_sheet.append(item)

     excel_sheet.column_dimensions[get_column_letter(1)].width=40

    excel_file.save(filename)
    excel_file.close()


placeNumber=input("정문=1, 중문=2, 서문=3, 후문=4")
my_place = input("지역을 입력하세요 : ")
food_name= input("음식을 입력하세요 : ")
file_name=input("저장할 파일 이름을 입력하세요 : ")

if placeNumber=='1':
    my_place="충북 청주시 흥덕구 복대동 680"
elif placeNumber=='2':
    my_place="충북 청주시 서원구 사창동 539"
elif placeNumber=='3':
    my_place="충북 청주시 서원구 성봉로242번길 31-25"
else:
    my_place="충북 청주시 서원구 1순환로 776"

driver = webdriver.Chrome()
driver.get("https://www.yogiyo.co.kr/")

driver.maximize_window()
time.sleep(1)

xpath = '''//*[@id="search"]/div/form/input'''
element = driver.find_element(By.XPATH, xpath)
element.clear()
time.sleep(1)

element.send_keys(my_place)
time.sleep(1)

search_button = '''//*[@id="button_search_address"]/button[2]'''
driver.find_element(By.XPATH, search_button).click()
time.sleep(1)

search_selector = '#search > div > form > ul > li:nth-child(3) > a'
search = driver.find_element(By.CSS_SELECTOR, search_selector)
search.click()
time.sleep(1)

input_box_xpath='''//*[@id="category"]/ul/li[1]/a'''
input_box=driver.find_element(By.XPATH, input_box_xpath)
time.sleep(1)
input_box.click()

send_food_xpath='''//*[@id="category"]/ul/li[15]/form/div/input'''
send_food=driver.find_element(By.XPATH, send_food_xpath)
time.sleep(1)
send_food.send_keys(food_name)

food_button_xpath='''//*[@id="category_search_button"]'''
food_button=driver.find_element(By.XPATH,food_button_xpath)
food_button.click()
time.sleep(1)

html = driver.page_source
html_source = BeautifulSoup(html, 'html.parser')

restaurantName = html_source.find_all("div", class_ = "restaurant-name ng-binding")
restaurantScore = html_source.find_all("span", class_ = "ico-star1 ng-binding") 
restaurantReview = html_source.find_all("span", attrs = {"class":"review_num ng-binding", "ng-show":"restaurant.review_count > 0"})
deliveryTime = html_source.find_all("li", class_ = "delivery-time ng-binding") 

data_list = []
profile_list = []

for i, j, k, m in zip(restaurantName, restaurantScore, restaurantReview, deliveryTime):
    data_list.append(i.string) 
    data_list.append(j.string.replace("★ ","")) 
    data_list.append(re.sub(" |\n|리뷰","",k.string)) 
    data_list.append(m.string.replace("\n","").replace(" ","")) 
    profile_list.append(data_list) 
    data_list = [] 
 
time.sleep(30)

write_excel_template(file_name, 'stores profile', profile_list)

driver.close() 
